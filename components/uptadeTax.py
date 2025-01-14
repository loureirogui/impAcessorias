import traceback
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import unicodedata
import re
import os
from art import *
from webdriver_manager.microsoft import EdgeChromiumDriverManager

def update_tax_regime(client_name, login_email, login_password, spreadsheet_name):

    def log_error(client_name, message):
        """
        Registra erros em um arquivo de texto dentro de uma pasta específica do cliente.
        """
        client_folder = f"./{client_name}"  # Cria o nome do diretório do cliente
        if not os.path.exists(client_folder):  # Cria a pasta, caso não exista
            os.makedirs(client_folder)
        
        error_log = os.path.join(client_folder, f"errors_{client_name}_regime.txt")  # Caminho do arquivo de erro
        with open(error_log, "a", encoding="utf-8") as file:  # Adiciona a mensagem de erro no arquivo
            file.write(message + "\n")

    def save_regime_log(client_name, regime_name, regime_id):
        """
        Salva informações de ações realizadas, como criação de regimes, em um arquivo de texto.
        """
        client_folder = f"./{client_name}"  # Cria o nome do diretório do cliente
        if not os.path.exists(client_folder):  # Cria a pasta, caso não exista
            os.makedirs(client_folder)

        log_path = os.path.join(client_folder, f"regime_logs_{client_name}.txt")  # Caminho do arquivo de log
        with open(log_path, "a", encoding="utf-8") as file:  # Adiciona o registro no arquivo
            file.write(f"{regime_name} - ID: {regime_id}\n")

    # Configuração inicial do navegador Edge
    edge_options = Options()
    edge_options.headless = False  # Define o modo headless como desativado
    browser = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()), options=edge_options)

    # Abre a página de login
    login_url = 'https://app.acessorias.com/sysmain.php?m=22'
    browser.get(login_url)

    # Realiza o login na aplicação
    try:
        email_field = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.NAME, 'mailAC'))
        )
        email_field.send_keys(login_email)
    except Exception:
        print("Erro ao inserir o e-mail no campo de login.")

    try:
        password_field = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.NAME, 'passAC'))
        )
        password_field.send_keys(login_password)
    except Exception:
        print("Erro ao inserir a senha no campo de senha.")

    try:
        login_button = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.button.rounded.large.expanded.primary-degrade.btn-enviar'))
        )
        login_button.click()
    except Exception:
        print("Erro ao clicar no botão de login.")

    # Espera o login ser concluído e abre a URL inicial
    time.sleep(2)
    browser.get(login_url)

    # Carrega a planilha Excel
    workbook = openpyxl.load_workbook(spreadsheet_name)
    sheet = workbook['Regime tributário']

    # Inativa regimes pré-definidos na aplicação
    try:
        predefined_regimes = WebDriverWait(browser, 10).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.dRow, .dOdd'))
        )
        for predefined_regime in predefined_regimes:
            browser.get(login_url)
            regime_links = WebDriverWait(browser, 10).until(
                EC.visibility_of_all_elements_located((By.XPATH, '//*[@id="main-container"]/div[1]/div[2]/div/div/div[4]/div[1]/a'))
            )
            for regime in regime_links:
                regime.click()  # Clica em cada regime
                try:
                    inactivate_field = WebDriverWait(browser, 10).until(
                        EC.visibility_of_element_located((By.NAME, 'RegAtivo'))
                    )
                    select = Select(inactivate_field)
                    select.select_by_visible_text('Não')  # Seleciona "Não" para inativar o regime
                except Exception as e:
                    print("Erro ao selecionar 'Não' para inativação.")
                    traceback.print_exc()
                browser.execute_script('check_form(this);')
                time.sleep(1)
                print("Regime tributário inativado com sucesso.")
    except Exception as e:
        print("Erro ao inativar regimes pré-definidos.")
        traceback.print_exc()

    # Função para normalizar texto (remover acentos e ajustar formato)
    def normalize_text(text):
        if text is None:
            return ""
        text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
        text = text.replace('º', '').replace('°', '').replace('ª', '').strip().lower()
        return text

    # Lê as colunas e cria novos regimes tributários
    column = 1
    while True:
        regime_name = sheet.cell(row=2, column=column).value
        if regime_name is None:  # Encerra quando encontra uma célula vazia
            break

        # Navega para a página de criação de novo regime
        creation_url = 'https://app.acessorias.com/sysmain.php?m=23&act=a&tr=R'
        browser.get(creation_url)

        try:
            regime_field = WebDriverWait(browser, 10).until(
                EC.visibility_of_element_located((By.NAME, 'RegNome'))
            )
            regime_field.send_keys(regime_name)
            print(f"Regime '{regime_name}' inserido com sucesso.")
            browser.execute_script('check_form(this);')# Simula o clique para salvar
            time.sleep(1)  
            log_error(client_name, f"Regime '{regime_name}' criado com sucesso.")

            # Adiciona obrigações ao regime
            row = 3
            empty_rows = 0
            while empty_rows < 5:
                obligation_name = sheet.cell(row=row, column=column).value
                if obligation_name is None:
                    empty_rows += 1
                else:
                    obligation_normalized = normalize_text(obligation_name)
                    empty_rows = 0

                    try:
                        obligation_selector = WebDriverWait(browser, 10).until(
                            EC.visibility_of_element_located((By.XPATH, '//*[@id="newObr"]'))
                        )
                        select = Select(obligation_selector)
                        for option in select.options:
                            if normalize_text(option.text).startswith(obligation_normalized):
                                select.select_by_visible_text(option.text)
                                add_button = WebDriverWait(browser, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="divSelectObr"]/button'))
                                )
                                ActionChains(browser).move_to_element(add_button).click().perform()
                                print(f"Obrigação '{obligation_name}' adicionada com sucesso.")
                                log_error(client_name, f"Obrigação '{obligation_name}' adicionada ao regime '{regime_name}'.")
                                break
                    except Exception as e:
                        print(f"Erro ao adicionar obrigação '{obligation_name}'.")
                        traceback.print_exc()

                row += 1

            browser.execute_script("check_form(this);")  # Salva o regime
            log_error(client_name, f"Regime '{regime_name}' salvo com sucesso.")
            time.sleep(2)

        except Exception as e:
            print(f"Erro ao criar ou salvar o regime '{regime_name}'.")
            log_error(client_name, f"Erro ao criar ou salvar o regime '{regime_name}'.")
        column += 1

    browser.quit()


